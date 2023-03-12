import shutil
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from  email.mime.base import MIMEBase
from email import encoders
import openpyxl
from openpyxl import load_workbook
"""
Before start script, always remove file 'Final_stat_form Счет-фактура № xxx от xx.xx.xxxx г..xlsx' from folder!
"""
class Country:
    def __init__(self,short_name_rus,code,alfa_2,short_name_en):
        self.short_name_rus = short_name_rus
        self.code = code
        self.alfa_2 = alfa_2
        self.short_name_en = short_name_en
    def __repr__(self):
        return self.short_name_rus + ' ' + self.code + ' ' + self.alfa_2 + ' ' + self.short_name_en

class Tnvd:
    def __init__(self,discription, tnvd_code):
        self.discription = discription
        self.tnvd_code = tnvd_code
def main():
    getInfoFromSF()
    sendMesageToMail()

def getInfoFromSF():
    doc_currency = str(input('Валюта документа рубли? y/n or/или да/нет'))
    rub_eur = float(input('Введите курс RUB к ЕВРО '))
    rub_usd = float(input('Введите курс RUB к USD '))
    lable = str(input('Введите бренд/торговую марку товара, например: Denso '))
    country_dict = {}
    tnvd_dict = {}

    file_tnvd = open(f'TNVD-list_{lable}.csv', 'r')
    for line in file_tnvd:
        read_line_tnvd_list = line.split(';')
        tmp_tnvd = Tnvd(read_line_tnvd_list[0].replace('"', ''), read_line_tnvd_list[1].replace('"', ''))
        tnvd_dict[tmp_tnvd.discription] = tmp_tnvd
    file_tnvd.close()

    file = open("country.csv", 'r')
    for line in file:
        read_line_list = line.split(';')
        tmp_country = Country(read_line_list[0].replace('"', ''), read_line_list[3].replace('"', ''),
                              read_line_list[4].replace('"', ''), read_line_list[6].replace('"', ''))
        country_dict[tmp_country.short_name_en] = tmp_country

    book = openpyxl.open('SF.xlsx', read_only=True, data_only=True)
    sheet = book.active
    row = 16
    data_list = []
    temp_data_list = []
    sf_number = sheet[3][0].value

    for x in range(16,(sheet.max_row)-10):
        for column in range(1,18):
            temp_data_list.append(sheet[row][column].value)
        data_list.append(temp_data_list)
        temp_data_list = []
        row +=1
    data_list = data_list[3:]
    for x in range(len(data_list)):
        data_list[x][1] = str(data_list[x][1])
    book.close()
    shutil.copy('Stat_form.xlsx', f'Final_stat_form {sf_number}.xlsx')
    work_book = openpyxl.open(f'Final_stat_form {sf_number}.xlsx', read_only=False, data_only=True)
    work_book_sheet = work_book.active
    row = 0

    for x in range(19,(sheet.max_row)-10):
        for column in range(1,23):
            work_book_sheet[row + 2][1].value = data_list[row][0] + " " + str(data_list[row][1])
            work_book_sheet[row + 2][0].value = tnvd_dict[data_list[row][1]].tnvd_code
            work_book_sheet[row + 2][10].value = data_list[row][15]
            work_book_sheet[row + 2][12].value = country_dict[(data_list[row][16].upper())].short_name_rus
            work_book_sheet[row + 2][11].value = data_list[row][16]
            work_book_sheet[row + 2][13].value = data_list[row][3]
            work_book_sheet[row + 2][14].value = data_list[row][2]
            work_book_sheet[row + 2][15].value = '796'
            work_book_sheet[row + 2][16].value = round(data_list[row][9],2)
            if doc_currency == 'y' or doc_currency == 'да':
                work_book_sheet[row + 2][17].value = round(data_list[row][9] / rub_usd, 2)
            else:
                work_book_sheet[row + 2][17].value = round(((data_list[row][9]) * rub_eur)/rub_usd,2)
            work_book_sheet[row + 2][18].value = round(data_list[row][9] * rub_eur, 2)
            work_book_sheet[row + 2][20].value = data_list[row][13]
        row +=1
    work_book.save(f'Final_stat_form {sf_number}.xlsx')
    work_book.close()


def sendMesageToMail():
    server = smtplib.SMTP('smtp.gmail.com', 587)
    sender = 'tableopposite@gmail.com'
    send_to = 'purchase2@bilight.biz, purchase@bilight.biz'
    password = 'rjonwqpruhzmdpte'
    server.starttls()
    message = 'Файл стат. декларирования готов для загрузки на таможню, но проверить не помешает:)'
    msg = MIMEMultipart()
    msg['From'] = sender
    msg['To'] = send_to
    msg['Subject'] = f'Файл для загрузки на ФТС {sf_number}'
    msg.attach(MIMEText(message))
    try:
        file = open(f'Final_stat_form {sf_number}.xlsx', 'rb')
        part = MIMEBase('application', f'Final_stat_form {sf_number}.xlsx')
        part.set_payload(file.read())
        file.close()
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename=f'Final_stat_form {sf_number}.xlsx')
        msg.attach(part)
        server.login(sender, password)
        server.sendmail(sender,send_to, msg.as_string())
        return print('Письмо отправленно успешно')
    except Exception as _ex:
        return f'{_ex}\n Проверьте ваш логин или пароль!'
    server.quit()

if __name__ == '__main__':
    main()