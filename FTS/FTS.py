import shutil
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from  email.mime.base import MIMEBase
from email import encoders
import openpyxl
from openpyxl import load_workbook
class Country:
    def __repr__(self):
        return self.code + ' ' + self.alfa_2 + ' ' + self.short_name

    def __init__(self,code,short_name,alfa_2):
        self.code = code
        self.short_name = short_name
        self.alfa_2 = alfa_2
class Row:
    def __init__(self,row):
        self.row = row
class Tnvd:
    def __init__(self,discription, tnvd_code):
        self.discription = discription
        self.tnvd_code = tnvd_code
def main():
    getInfoFromSF()
    sendMesageToMail()

def getInfoFromSF():

    amd = float(input('Введите курс AMD '))
    usd = float(input('Введите курс USD '))
    country_of_origin = (str(input('Введите страну происхождения товара заглавными буквами например: АРМЕНИЯ '))).upper()
    lable = ''
    lable = str(input('Введите бренд/торговую марку товара, например: Denso '))
    country_dict = {}
    tnvd_dict = {}
    file_tnvd = open(f'TNVD-list_{lable}.csv', 'r')
    for line in file_tnvd:
        read_line_tnvd_list = line.split(';')
        tmp_tnvd = Tnvd(read_line_tnvd_list[0].replace('"', ''), read_line_tnvd_list[1].replace('"', ''))
        tnvd_dict[tmp_tnvd.discription] = tmp_tnvd
    file_tnvd.close()
    file = open("country.csv", 'r')
    for line in file:
        read_line_list = line.split(';')
        tmp_country = Country(read_line_list[3].replace('"', ''), read_line_list[0].replace('"', ''),
                              read_line_list[4].replace('"', ''))
        country_dict[tmp_country.short_name] = tmp_country
    book = openpyxl.open('SF.xlsx', read_only=True, data_only=True)
    sheet = book.active
    row = Row(16)
    data_list = []
    temp_data_list = []
    for x in range(16,(sheet.max_row)-6):
        for column in range(1,14):
            temp_data_list.append(sheet[row.row][column].value)
        data_list.append(temp_data_list)
        temp_data_list = []
        row.row +=1
    book.close()
    weight_book = openpyxl.open('Weight.xlsx', read_only=True, data_only=True)
    sheet_weight = weight_book.active
    weight_list = []
    for x in range(16, (sheet_weight.max_row) - 6):
        weight_list.append(sheet_weight[x][14].value)
    weight_book.close()
    shutil.copy('Stat_form.xlsx', 'Final_stat_form.xlsx')
    work_book = openpyxl.open('Final_stat_form.xlsx', read_only=False, data_only=True)
    work_book_sheet = work_book.active
    row.row = 0
    print(data_list)
    for x in range(16,(sheet.max_row)-6):
        for column in range(1,23):
            work_book_sheet[row.row + 2][1].value = data_list[row.row][1]
            work_book_sheet[row.row + 2][0].value = tnvd_dict[data_list[row.row][2]].tnvd_code
            work_book_sheet[row.row + 2][10].value = weight_list[row.row]
            work_book_sheet[row.row + 2][12].value = country_of_origin
            work_book_sheet[row.row + 2][11].value = country_dict[country_of_origin].alfa_2
            work_book_sheet[row.row + 2][13].value = data_list[row.row][6]
            work_book_sheet[row.row + 2][14].value = data_list[row.row][5]
            work_book_sheet[row.row + 2][15].value = '796'
            work_book_sheet[row.row + 2][16].value = data_list[row.row][12]
            work_book_sheet[row.row + 2][17].value = round((data_list[row.row][12]*amd)/usd,2)
            work_book_sheet[row.row + 2][18].value = round(data_list[row.row][12] * amd, 2)
        row.row +=1
    for x in range(0, (sheet.max_row)-22):
        if lable == 'Narva':
            work_book_sheet[x+2][1].value = work_book_sheet[x+2][1].value.replace('Lamp', 'Лампа автомобильная') + ' ' + 'арт. ' + str(data_list[x][2])
        else:
           work_book_sheet[x + 2][1].value = work_book_sheet[x+2][1].value.replace('WIPERBLADE','Щетка стеклоочистителя') + ' ' + 'арт. ' + str(data_list[x][2])
           work_book_sheet[x + 2][1].value = work_book_sheet[x + 2][1].value.replace('REAR Щетка', 'Задняя щетка')
           work_book_sheet[x + 2][1].value = work_book_sheet[x + 2][1].value.replace('FLAT BLADE RETROFIT','Щетка стеклоочистителя FLAT BLADE RETROFIT')
           work_book_sheet[x + 2][1].value = work_book_sheet[x + 2][1].value.replace('SPARK PLUG', 'Свеча зажигания')
    work_book.save('Final_stat_form.xlsx')
    work_book.close()
def sendMesageToMail():
    server = smtplib.SMTP('smtp.gmail.com', 587)
    sender = 'tableopposite@gmail.com'
    send_to = 'purchase2@bilight.biz, purchase@bilight.biz'
    password = 'rjonwqpruhzmdpte'
    server.starttls()
    message = 'Файл стат. декларирования готов для загрузки на таможню, но проверить не помешает:)'
    msg = MIMEMultipart()
    msg['From'] = sender
    msg['To'] = send_to
    msg['Subject'] = 'Файл для загрузки на ФТС'
    msg.attach(MIMEText(message))
    try:
        file = open('Final_stat_form.xlsx', 'rb')
        part = MIMEBase('application', 'Final_stat_form.xlsx')
        part.set_payload(file.read())
        file.close()
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename='Final_stat_form.xlsx')
        msg.attach(part)
        server.login(sender, password)
        server.sendmail(sender,send_to, msg.as_string())
        return print('Письмо отправленно успешно')
    except Exception as _ex:
        return f'{_ex}\n Проверьте ваш логин или пароль!'
    server.quit()

if __name__ == '__main__':
    main()